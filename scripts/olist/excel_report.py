# -*- coding: utf-8 -*-
"""
作品二 · Olist 业务汇报 Excel
====================================================
把 pandas 分析结果导出为多 sheet 中文业务报表(供领导/跨部门查看)。
Sheet:
1. 客户概况  - 整体指标
2. 消费分层  - 4 档价值分层
3. 复购对比  - 复购 vs 单次
4. 运营靶心  - 高价值单次客户
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 读取分析结果
客户汇总 = pd.read_parquet("output/olist/customer_summary.parquet")
客户汇总["价值档"] = pd.qcut(客户汇总["总消费"], 4, labels=["低", "中低", "中高", "高"])

# 计算核心指标
复购客户 = 客户汇总[客户汇总["订单数"] >= 2]
单次客户 = 客户汇总[客户汇总["订单数"] == 1]
阈值 = 单次客户["总消费"].quantile(0.8)
高价值单次 = 单次客户[单次客户["总消费"] >= 阈值]

# ========== 样式工具 ==========
HEADER_FILL = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1E40AF")
THIN = Side(border_style="thin", color="DBEAFE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def write_table(ws, start_row, headers, rows):
    """写入表头+数据,返回结束行号"""
    # 表头
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER
    # 数据
    for r, row in enumerate(rows, start_row + 1):
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal="center" if isinstance(v, (int, float)) else "left")
    return start_row + len(rows)

wb = Workbook()

# ========== Sheet 1: 客户概况 ==========
ws = wb.active
ws.title = "客户概况"
ws["A1"] = "Olist 电商客户分析 · 业务汇报"
ws["A1"].font = TITLE_FONT
概况 = [
    ("有效订单数", "99,440"),
    ("真实客户数", f"{len(客户汇总):,}"),
    ("平均消费(R$)", f"{客户汇总['总消费'].mean():.1f}"),
    ("消费中位数(R$)", f"{客户汇总['总消费'].median():.1f}"),
    ("只买1单客户占比", f"{(客户汇总['订单数']==1).mean():.1%}"),
    ("复购客户数", f"{len(复购客户):,}"),
    ("复购客户平均消费(R$)", f"{复购客户['总消费'].mean():.1f}"),
    ("单次客户平均消费(R$)", f"{单次客户['总消费'].mean():.1f}"),
    ("复购/单次价值倍数", f"{复购客户['总消费'].mean()/单次客户['总消费'].mean():.2f}x"),
    ("高价值单次客户数", f"{len(高价值单次):,}"),
    ("高价值单次转化阈值(R$)", f"{阈值:.1f}"),
]
ws.cell(row=3, column=1, value="指标").font = HEADER_FONT
ws.cell(row=3, column=2, value="数值").font = HEADER_FONT
for i, (k, v) in enumerate(概况, 4):
    ws.cell(row=i, column=1, value=k).border = BORDER
    ws.cell(row=i, column=2, value=v).border = BORDER
    ws.cell(row=i, column=1).font = Font(bold=True)
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 20

# ========== Sheet 2: 消费分层 ==========
ws2 = wb.create_sheet("消费分层")
ws2["A1"] = "消费分位数 4 档价值分层"
ws2["A1"].font = TITLE_FONT
分层 = 客户汇总.groupby("价值档", observed=True).agg(
    人数=("总消费", "count"),
    平均消费=("总消费", "mean"),
    平均订单数=("订单数", "mean"),
).round(1)
分层 = 分层.reindex(["低", "中低", "中高", "高"])
write_table(ws2, 3, ["价值档", "人数", "平均消费", "平均订单数"],
            [[idx, v["人数"], v["平均消费"], v["平均订单数"]] for idx, v in 分层.iterrows()])
for c in range(1, 5):
    ws2.column_dimensions[get_column_letter(c)].width = 16

# ========== Sheet 3: 复购对比 ==========
ws3 = wb.create_sheet("复购对比")
ws3["A1"] = "复购客户 vs 单次客户"
ws3["A1"].font = TITLE_FONT
复购对比 = [
    ("复购客户", len(复购客户), 复购客户["总消费"].mean()),
    ("单次客户", len(单次客户), 单次客户["总消费"].mean()),
]
write_table(ws3, 3, ["客户类型", "人数", "平均消费(R$)"],
            [[t, n, round(v, 1)] for t, n, v in 复购对比])
ws3.cell(row=8, column=1, value=f"复购价值是单次的 {复购客户['总消费'].mean()/单次客户['总消费'].mean():.2f} 倍").font = Font(bold=True, color="D97706")
for c in range(1, 4):
    ws3.column_dimensions[get_column_letter(c)].width = 16

# ========== Sheet 4: 运营靶心 ==========
ws4 = wb.create_sheet("运营靶心")
ws4["A1"] = "高价值单次客户(复购转化靶心)"
ws4["A1"].font = TITLE_FONT
靶心 = [
    ("单次客户总数", len(单次客户)),
    ("高价值单次客户(消费≥203)", len(高价值单次)),
    ("占单次客户比例", f"{len(高价值单次)/len(单次客户):.1%}"),
    ("高价值单次平均消费", f"{高价值单次['总消费'].mean():.1f}"),
]
write_table(ws4, 3, ["指标", "数值"], [[k, v] for k, v in 靶心])
ws4.cell(row=9, column=1, value="运营动作:定向优惠券 / 会员体系 / 购买后召回").font = Font(bold=True, color="D97706")
ws4.column_dimensions["A"].width = 40
ws4.column_dimensions["B"].width = 24

out = "output/olist/Olist_业务汇报.xlsx"
wb.save(out)
print(f"Excel 报表已生成: {out}")
print(f"Sheets: 客户概况 / 消费分层 / 复购对比 / 运营靶心")
